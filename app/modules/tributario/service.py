import csv
import io
import os
import re
import time
import unicodedata
from datetime import date, datetime, timezone
from typing import Any, Dict, List, Optional, Tuple

import requests
from fastapi import Depends, HTTPException
from fastapi.security import HTTPAuthorizationCredentials, HTTPBearer


# ============================================================
# CONFIGURAÇÕES
# ============================================================

# Campinas / São Paulo
TINY_TOKEN = os.getenv("TINY_TOKEN", "").strip()

# Pouso Alegre / Minas Gerais
TINY_API_KEY_MINAS = os.getenv("TINY_API_KEY_MINAS", "").strip()

TINY_BASE_URL = "https://api.tiny.com.br/api2"

SUPABASE_URL = os.getenv("SUPABASE_URL", "").strip().rstrip("/")
SUPABASE_SERVICE_ROLE_KEY = (
    os.getenv("SUPABASE_SERVICE_ROLE_KEY", "").strip()
    or os.getenv("SUPABASE_SERVICE_KEY", "").strip()
)

TABELA_PRODUTOS = "produtos_tributarios"

# Mantém o mesmo comportamento do main.py enviado.
JWT_AUTH_ENABLED = os.getenv("JWT_AUTH_ENABLED", "false").strip().lower() in {
    "1", "true", "yes", "sim", "on"
}

bearer_scheme = HTTPBearer(auto_error=False)

# Colunas que realmente existem na tabela produtos_tributarios.
COLUNAS_PRODUTOS_TRIBUTARIOS = {
    "id",
    "filial",
    "tiny_produto_id",
    "codigo",
    "sku",
    "descricao",
    "ncm",
    "cest",
    "origem_mercadoria",
    "unidade",
    "preco",
    "custo",
    "fornecedor",
    "gtin",
    "ativo",
    "ultima_sincronizacao",
    "created_at",
    "updated_at",
    "status_ia",
    "percentual_confianca",
    "justificativa_ia",
    "data_analise",
    "lote",
}

# Campos que podem ser alterados pelo endpoint PATCH.
# id, filial, tiny_produto_id e created_at não devem ser modificados manualmente.
COLUNAS_EDITAVEIS = {
    "codigo",
    "sku",
    "descricao",
    "ncm",
    "cest",
    "origem_mercadoria",
    "unidade",
    "preco",
    "custo",
    "fornecedor",
    "gtin",
    "ativo",
    "status_ia",
    "percentual_confianca",
    "justificativa_ia",
    "data_analise",
    "lote",
}


# ============================================================
# FUNÇÕES BÁSICAS
# ============================================================

def agora_iso() -> str:
    """Retorna data/hora UTC em ISO 8601."""
    return datetime.now(timezone.utc).isoformat()


def safe_str(valor: Any) -> str:
    return "" if valor is None else str(valor).strip()


def dinheiro_para_float(valor: Any) -> float:
    """
    Converte números e valores monetários brasileiros para float.

    Exemplos:
    10 -> 10.0
    "10,50" -> 10.5
    "R$ 1.234,56" -> 1234.56
    """
    if valor is None:
        return 0.0

    if isinstance(valor, bool):
        return float(valor)

    if isinstance(valor, (int, float)):
        return float(valor)

    texto = (
        str(valor)
        .replace("R$", "")
        .replace("\u00a0", "")
        .replace(" ", "")
        .strip()
    )

    if not texto:
        return 0.0

    if "," in texto and "." in texto:
        texto = texto.replace(".", "").replace(",", ".")
    elif "," in texto:
        texto = texto.replace(",", ".")

    try:
        return float(texto)
    except (TypeError, ValueError):
        return 0.0


def normalizar_filial(filial: str) -> str:
    """
    Normaliza o código da filial.

    sp  = Campinas
    mg  = Pouso Alegre
    all = consolidado, apenas para consultas
    """
    valor = safe_str(filial).lower()

    if valor in {"mg", "minas", "pouso_alegre", "pouso-alegre", "pouso alegre"}:
        return "mg"

    if valor in {"all", "todas", "todos", "consolidado"}:
        return "all"

    return "sp"


def validar_configuracao() -> None:
    faltando: List[str] = []

    if not SUPABASE_URL:
        faltando.append("SUPABASE_URL")

    if not SUPABASE_SERVICE_ROLE_KEY:
        faltando.append("SUPABASE_SERVICE_ROLE_KEY ou SUPABASE_SERVICE_KEY")

    if faltando:
        raise HTTPException(
            status_code=500,
            detail={
                "erro": "Variáveis de ambiente ausentes.",
                "faltando": faltando,
            },
        )


def obter_token_tiny(filial: str) -> str:
    """
    Seleciona o Tiny correto sem misturar as filiais.

    filial=sp -> TINY_TOKEN
    filial=mg -> TINY_API_KEY_MINAS
    """
    filial_normalizada = normalizar_filial(filial)

    if filial_normalizada == "all":
        raise HTTPException(
            status_code=400,
            detail="Para sincronizar, use filial=sp ou filial=mg.",
        )

    if filial_normalizada == "mg":
        token = TINY_API_KEY_MINAS
        nome_variavel = "TINY_API_KEY_MINAS"
    else:
        token = TINY_TOKEN
        nome_variavel = "TINY_TOKEN"

    if not token:
        raise HTTPException(
            status_code=500,
            detail=f"{nome_variavel} não configurado.",
        )

    return token


def ncm_limpo(valor: Any) -> str:
    """Mantém apenas os números do NCM."""
    return "".join(caractere for caractere in safe_str(valor) if caractere.isdigit())


def status_ncm(ncm: Any) -> str:
    """
    Status tributário calculado em memória.

    Este status NÃO é salvo porque a tabela atual não possui
    a coluna status_tributario.
    """
    valor = ncm_limpo(ncm)

    if not valor:
        return "sem_ncm"

    if len(valor) != 8:
        return "ncm_invalido"

    return "pendente"


def status_ia_inicial(ncm: Any) -> str:
    valor = ncm_limpo(ncm)

    if not valor:
        return "PENDENTE"

    if len(valor) != 8:
        return "REVISAO_MANUAL"

    return "PENDENTE"


def adicionar_status_tributario_virtual(
    registro: Dict[str, Any],
) -> Dict[str, Any]:
    """
    Acrescenta status_tributario somente na resposta da API.

    Assim o frontend continua recebendo o campo, mas o Supabase
    não recebe uma coluna inexistente.
    """
    resultado = dict(registro)
    resultado["status_tributario"] = status_ncm(resultado.get("ncm"))
    return resultado



# ============================================================
# AUTENTICAÇÃO COMPATÍVEL COM O MAIN.PY
# ============================================================

def validar_token_supabase(access_token: str) -> Dict[str, Any]:
    """Valida o Bearer token no Supabase Auth, igual ao main.py."""
    validar_configuracao()

    response = requests.get(
        f"{SUPABASE_URL}/auth/v1/user",
        headers={
            "apikey": SUPABASE_SERVICE_ROLE_KEY,
            "Authorization": f"Bearer {access_token}",
            "Content-Type": "application/json",
        },
        timeout=30,
    )

    if response.status_code in {401, 403}:
        raise HTTPException(
            status_code=401,
            detail="Sessão inválida ou expirada. Faça login novamente.",
        )

    if response.status_code >= 400:
        raise HTTPException(
            status_code=503,
            detail={
                "erro": "Não foi possível validar a sessão no Supabase Auth.",
                "status_code": response.status_code,
                "resposta": response.text,
            },
        )

    usuario = response.json()
    if not usuario.get("id") or not usuario.get("email"):
        raise HTTPException(
            status_code=401,
            detail="Token válido, mas sem identificação de usuário.",
        )
    return usuario


def buscar_perfil_dashboard(email: str) -> Dict[str, Any]:
    """Busca a autorização interna na tabela usuarios_dashboard."""
    response = requests.get(
        f"{SUPABASE_URL}/rest/v1/usuarios_dashboard",
        headers=supabase_headers(),
        params={
            "select": "email,nome,perfil,filial,ativo",
            "email": f"eq.{email}",
            "limit": "1",
        },
        timeout=60,
    )

    if response.status_code >= 400:
        raise HTTPException(
            status_code=500,
            detail={
                "erro": "Erro ao consultar usuarios_dashboard.",
                "status_code": response.status_code,
                "resposta": response.text,
            },
        )

    dados = response.json()
    if not dados:
        raise HTTPException(
            status_code=403,
            detail="Usuário autenticado, mas sem acesso ao Dashboard MHM.",
        )

    perfil = dados[0]
    if perfil.get("ativo") is False:
        raise HTTPException(
            status_code=403,
            detail="Usuário desativado no Dashboard MHM.",
        )

    perfil["perfil"] = safe_str(perfil.get("perfil")).lower()
    perfil["filial"] = normalizar_filial(perfil.get("filial") or "sp")
    return perfil


def obter_usuario_atual(
    credenciais: Optional[HTTPAuthorizationCredentials] = Depends(bearer_scheme),
) -> Dict[str, Any]:
    """Dependência de autenticação com o mesmo modo gradual do main.py."""
    if not JWT_AUTH_ENABLED:
        return {
            "id": "modo-implantacao",
            "email": "sistema@interno",
            "nome": "Modo de implantação",
            "perfil": "admin",
            "filial": "all",
            "ativo": True,
            "auth_desativada_temporariamente": True,
        }

    if not credenciais or not credenciais.credentials:
        raise HTTPException(
            status_code=401,
            detail="Token de acesso não informado.",
            headers={"WWW-Authenticate": "Bearer"},
        )

    usuario_auth = validar_token_supabase(credenciais.credentials)
    perfil = buscar_perfil_dashboard(usuario_auth["email"])
    return {
        "id": usuario_auth["id"],
        "email": usuario_auth["email"],
        "nome": perfil.get("nome") or usuario_auth["email"],
        "perfil": perfil.get("perfil"),
        "filial": perfil.get("filial"),
        "ativo": perfil.get("ativo", True),
    }


def resolver_filial_autorizada(
    filial_solicitada: Optional[str],
    usuario: Dict[str, Any],
    permitir_all: bool = True,
) -> str:
    """Aplica a mesma separação SP/MG usada no main.py."""
    filial_usuario = normalizar_filial(usuario.get("filial") or "sp")

    if filial_solicitada is None or not safe_str(filial_solicitada):
        if filial_usuario == "all":
            return "all" if permitir_all else "sp"
        return filial_usuario

    filial_pedida = normalizar_filial(filial_solicitada)

    if not permitir_all and filial_pedida == "all":
        raise HTTPException(
            status_code=400,
            detail="Esta operação não aceita filial=all.",
        )

    if filial_usuario == "all":
        return filial_pedida

    if filial_pedida != filial_usuario:
        raise HTTPException(
            status_code=403,
            detail=f"Usuário sem permissão para acessar a filial '{filial_pedida}'.",
        )

    return filial_usuario


# ============================================================
# SUPABASE
# ============================================================

def supabase_headers(prefer: Optional[str] = None) -> Dict[str, str]:
    validar_configuracao()

    headers = {
        "apikey": SUPABASE_SERVICE_ROLE_KEY,
        "Authorization": f"Bearer {SUPABASE_SERVICE_ROLE_KEY}",
        "Content-Type": "application/json",
    }

    if prefer:
        headers["Prefer"] = prefer

    return headers


def supabase_get(params: Any) -> List[Dict[str, Any]]:
    response = requests.get(
        f"{SUPABASE_URL}/rest/v1/{TABELA_PRODUTOS}",
        headers=supabase_headers(),
        params=params,
        timeout=90,
    )

    if response.status_code >= 400:
        raise HTTPException(
            status_code=500,
            detail={
                "erro": f"Erro ao consultar {TABELA_PRODUTOS}.",
                "status_code": response.status_code,
                "resposta": response.text,
            },
        )

    try:
        dados = response.json()
    except ValueError:
        raise HTTPException(
            status_code=500,
            detail={
                "erro": "Supabase não retornou JSON válido.",
                "resposta": response.text,
            },
        )

    return dados if isinstance(dados, list) else []


def supabase_get_todos(
    params_base: Any,
    tamanho_pagina: int = 1000,
    limite_total: Optional[int] = None,
) -> List[Dict[str, Any]]:
    """
    Lê todos os registros de produtos_tributarios usando paginação.

    O Supabase pode limitar cada resposta a 100 ou 1.000 linhas, mesmo
    quando o parâmetro limit é maior. Por isso, avançamos pelo número
    real de registros recebidos até a página ficar vazia.
    """
    if tamanho_pagina < 1 or tamanho_pagina > 10000:
        raise HTTPException(
            status_code=400,
            detail="tamanho_pagina deve estar entre 1 e 10000.",
        )

    params_fixos = [
        (str(chave), str(valor))
        for chave, valor in list(params_base)
        if str(chave) not in {"limit", "offset"}
    ]

    registros: List[Dict[str, Any]] = []
    offset = 0

    while True:
        restante = None
        if limite_total is not None:
            restante = limite_total - len(registros)
            if restante <= 0:
                break

        tamanho_solicitado = (
            min(tamanho_pagina, restante)
            if restante is not None
            else tamanho_pagina
        )

        params = [
            *params_fixos,
            ("limit", str(tamanho_solicitado)),
            ("offset", str(offset)),
        ]
        pagina = supabase_get(params)

        if not pagina:
            break

        registros.extend(pagina)
        offset += len(pagina)

        # Quando o limite total foi atingido, encerramos. Não usamos
        # len(pagina) < tamanho_solicitado como condição porque o projeto
        # Supabase pode impor um teto menor por resposta.
        if limite_total is not None and len(registros) >= limite_total:
            break

    return registros


def limpar_payload_banco(dados: Dict[str, Any]) -> Dict[str, Any]:
    """
    Remove campos que não existem na tabela.

    Também converte nomes antigos para os nomes reais:
    preco_venda -> preco
    preco_custo -> custo
    """
    origem = dict(dados)

    if "preco" not in origem and "preco_venda" in origem:
        origem["preco"] = origem.get("preco_venda")

    if "custo" not in origem and "preco_custo" in origem:
        origem["custo"] = origem.get("preco_custo")

    origem.pop("preco_venda", None)
    origem.pop("preco_custo", None)
    origem.pop("status_tributario", None)

    return {
        chave: valor
        for chave, valor in origem.items()
        if chave in COLUNAS_PRODUTOS_TRIBUTARIOS
    }


def supabase_upsert(payload: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    if not payload:
        return []

    payload_limpo = [limpar_payload_banco(item) for item in payload]

    response = requests.post(
        f"{SUPABASE_URL}/rest/v1/{TABELA_PRODUTOS}",
        headers=supabase_headers(
            "resolution=merge-duplicates,return=representation"
        ),
        params={"on_conflict": "filial,tiny_produto_id"},
        json=payload_limpo,
        timeout=120,
    )

    if response.status_code >= 400:
        raise HTTPException(
            status_code=500,
            detail={
                "erro": f"Erro ao salvar em {TABELA_PRODUTOS}.",
                "status_code": response.status_code,
                "resposta": response.text,
            },
        )

    try:
        dados = response.json()
    except ValueError:
        return []

    return dados if isinstance(dados, list) else []


def supabase_patch(
    produto_id: str,
    dados: Dict[str, Any],
) -> List[Dict[str, Any]]:
    """
    Atualiza um produto sem permitir o envio de colunas inexistentes.
    """
    dados_normalizados = limpar_payload_banco(dados)

    dados_editaveis = {
        chave: valor
        for chave, valor in dados_normalizados.items()
        if chave in COLUNAS_EDITAVEIS
    }

    if not dados_editaveis:
        raise HTTPException(
            status_code=400,
            detail="Nenhum campo válido foi enviado para atualização.",
        )

    # Se o NCM foi alterado e o status da IA não veio no corpo,
    # recalcula o estado inicial.
    if "ncm" in dados_editaveis:
        ncm = ncm_limpo(dados_editaveis.get("ncm"))
        dados_editaveis["ncm"] = ncm or None

        if "status_ia" not in dados_editaveis:
            dados_editaveis["status_ia"] = status_ia_inicial(ncm)

    dados_editaveis["updated_at"] = agora_iso()

    response = requests.patch(
        f"{SUPABASE_URL}/rest/v1/{TABELA_PRODUTOS}",
        headers=supabase_headers("return=representation"),
        params={"id": f"eq.{produto_id}"},
        json=dados_editaveis,
        timeout=90,
    )

    if response.status_code >= 400:
        raise HTTPException(
            status_code=500,
            detail={
                "erro": "Erro ao atualizar produto tributário.",
                "status_code": response.status_code,
                "resposta": response.text,
            },
        )

    try:
        registros = response.json()
    except ValueError:
        registros = []

    return [
        adicionar_status_tributario_virtual(registro)
        for registro in registros
    ]


# ============================================================
# TINY
# ============================================================

def tiny_get(
    endpoint: str,
    params: Dict[str, Any],
    filial: str,
) -> Dict[str, Any]:
    parametros = {
        "token": obter_token_tiny(filial),
        "formato": "json",
        **params,
    }

    response = requests.get(
        f"{TINY_BASE_URL}/{endpoint}",
        params=parametros,
        timeout=90,
    )

    if response.status_code >= 400:
        raise HTTPException(
            status_code=502,
            detail={
                "erro": "Erro HTTP ao consultar o Tiny.",
                "status_code": response.status_code,
                "resposta": response.text,
            },
        )

    try:
        dados = response.json()
    except ValueError:
        raise HTTPException(
            status_code=502,
            detail={
                "erro": "Tiny não retornou JSON válido.",
                "resposta": response.text,
            },
        )

    retorno = dados.get("retorno", {})

    if safe_str(retorno.get("status")).upper() == "ERRO":
        codigo = safe_str(retorno.get("codigo_erro"))
        erros = retorno.get("erros") or []

        texto = " ".join(
            safe_str(item.get("erro") if isinstance(item, dict) else item)
            for item in erros
        ).lower()

        # O Tiny usa esse erro quando a página não possui registros.
        if (
            codigo == "20"
            or "não retornou registros" in texto
            or "nao retornou registros" in texto
        ):
            return {
                "retorno": {
                    "status": "OK",
                    "numero_paginas": 1,
                    "produtos": [],
                }
            }

        raise HTTPException(
            status_code=502,
            detail={
                "erro": "Tiny retornou erro.",
                "retorno": retorno,
            },
        )

    return dados


def buscar_pagina_produtos_tiny(
    pagina: int,
    filial: str,
) -> Dict[str, Any]:
    return tiny_get(
        "produtos.pesquisa.php",
        {
            "pagina": pagina,
            "situacao": "A",
        },
        filial,
    )


def obter_produto_tiny(
    produto_id: str,
    filial: str,
) -> Dict[str, Any]:
    dados = tiny_get(
        "produto.obter.php",
        {"id": produto_id},
        filial,
    )

    produto = dados.get("retorno", {}).get("produto", {})
    return produto if isinstance(produto, dict) else {}


def extrair_produtos_pagina(
    resposta: Dict[str, Any],
) -> Tuple[List[Dict[str, Any]], int]:
    retorno = resposta.get("retorno", {})
    produtos: List[Dict[str, Any]] = []

    for item in retorno.get("produtos", []) or []:
        if not isinstance(item, dict):
            continue

        produto = item.get("produto", item)

        if isinstance(produto, dict):
            produtos.append(produto)

    try:
        total_paginas = int(retorno.get("numero_paginas", 1) or 1)
    except (TypeError, ValueError):
        total_paginas = 1

    return produtos, max(total_paginas, 1)


# ============================================================
# NORMALIZAÇÃO DO PRODUTO
# ============================================================

def extrair_fornecedor(detalhe: Dict[str, Any]) -> Optional[str]:
    fornecedor = detalhe.get("fornecedor") or {}

    if isinstance(fornecedor, dict):
        nome = safe_str(
            fornecedor.get("nome")
            or fornecedor.get("nome_fantasia")
            or fornecedor.get("razao_social")
        )
    else:
        nome = safe_str(fornecedor)

    return nome or None


def normalizar_produto(
    produto_pesquisa: Dict[str, Any],
    filial: str,
    lote: int,
    produto_detalhado: Optional[Dict[str, Any]] = None,
) -> Dict[str, Any]:
    detalhe = produto_detalhado or {}

    produto_id = safe_str(
        detalhe.get("id")
        or produto_pesquisa.get("id")
    )

    codigo = safe_str(
        detalhe.get("codigo")
        or produto_pesquisa.get("codigo")
    )

    descricao = safe_str(
        detalhe.get("nome")
        or detalhe.get("descricao")
        or produto_pesquisa.get("nome")
        or produto_pesquisa.get("descricao")
    )

    ncm = ncm_limpo(
        detalhe.get("ncm")
        or produto_pesquisa.get("ncm")
    )

    cest = safe_str(
        detalhe.get("cest")
        or produto_pesquisa.get("cest")
    )

    custo = dinheiro_para_float(
        detalhe.get("preco_custo")
        or detalhe.get("preco_custo_medio")
        or detalhe.get("custo")
        or produto_pesquisa.get("preco_custo")
        or produto_pesquisa.get("custo")
    )

    preco = dinheiro_para_float(
        detalhe.get("preco")
        or detalhe.get("preco_venda")
        or produto_pesquisa.get("preco")
        or produto_pesquisa.get("preco_venda")
    )

    agora = agora_iso()

    # ATENÇÃO:
    # O retorno abaixo usa somente colunas existentes no Supabase.
    # Não envia preco_custo, preco_venda nem status_tributario.
    return {
        "filial": normalizar_filial(filial),
        "tiny_produto_id": produto_id,
        "codigo": codigo or None,
        "sku": codigo or None,
        "descricao": descricao or None,
        "ncm": ncm or None,
        "cest": cest or None,
        "origem_mercadoria": (
            safe_str(
                detalhe.get("origem")
                or detalhe.get("origem_mercadoria")
                or produto_pesquisa.get("origem")
                or produto_pesquisa.get("origem_mercadoria")
            )
            or None
        ),
        "unidade": (
            safe_str(
                detalhe.get("unidade")
                or produto_pesquisa.get("unidade")
            )
            or None
        ),
        "preco": preco,
        "custo": custo,
        "fornecedor": extrair_fornecedor(detalhe),
        "ativo": True,
        "status_ia": status_ia_inicial(ncm),
        "lote": lote,
        "ultima_sincronizacao": agora,
        "updated_at": agora,
    }


# ============================================================
# SINCRONIZAÇÃO
# ============================================================

def sincronizar_produtos(
    filial: str,
    tamanho_lote: int = 50,
    detalhar: bool = True,
    limite_detalhes: int = 50,
    max_paginas: Optional[int] = None,
) -> Dict[str, Any]:
    filial_normalizada = normalizar_filial(filial)

    if filial_normalizada == "all":
        raise HTTPException(
            status_code=400,
            detail="Use filial=sp ou filial=mg.",
        )

    if tamanho_lote < 1 or tamanho_lote > 500:
        raise HTTPException(
            status_code=400,
            detail="tamanho_lote deve estar entre 1 e 500.",
        )

    if limite_detalhes < 0:
        raise HTTPException(
            status_code=400,
            detail="limite_detalhes não pode ser negativo.",
        )

    if max_paginas is not None and max_paginas < 1:
        raise HTTPException(
            status_code=400,
            detail="max_paginas deve ser maior ou igual a 1.",
        )

    pagina = 1
    paginas_processadas = 0
    indice_global = 0
    detalhes_executados = 0
    produtos_recebidos = 0
    produtos_salvos = 0
    produtos_ignorados_sem_id = 0
    sem_ncm = 0
    ncm_invalido = 0
    erros_detalhe: List[Dict[str, str]] = []

    while True:
        resposta = buscar_pagina_produtos_tiny(
            pagina,
            filial_normalizada,
        )

        produtos_pagina, total_paginas = extrair_produtos_pagina(resposta)
        paginas_processadas += 1
        produtos_recebidos += len(produtos_pagina)

        payload: List[Dict[str, Any]] = []

        for produto in produtos_pagina:
            indice_global += 1
            numero_lote = ((indice_global - 1) // tamanho_lote) + 1

            detalhe: Optional[Dict[str, Any]] = None
            produto_id = safe_str(produto.get("id"))

            precisa_detalhar = (
                detalhar
                and bool(produto_id)
                and detalhes_executados < limite_detalhes
            )

            if precisa_detalhar:
                try:
                    detalhe = obter_produto_tiny(
                        produto_id,
                        filial_normalizada,
                    )
                    detalhes_executados += 1
                    time.sleep(0.55)

                except HTTPException as exc:
                    erros_detalhe.append(
                        {
                            "tiny_produto_id": produto_id,
                            "erro": safe_str(exc.detail)[:300],
                        }
                    )

                except Exception as exc:
                    erros_detalhe.append(
                        {
                            "tiny_produto_id": produto_id,
                            "erro": str(exc)[:300],
                        }
                    )

            normalizado = normalizar_produto(
                produto_pesquisa=produto,
                filial=filial_normalizada,
                lote=numero_lote,
                produto_detalhado=detalhe,
            )

            if not normalizado.get("tiny_produto_id"):
                produtos_ignorados_sem_id += 1
                continue

            status_calculado = status_ncm(normalizado.get("ncm"))

            if status_calculado == "sem_ncm":
                sem_ncm += 1
            elif status_calculado == "ncm_invalido":
                ncm_invalido += 1

            payload.append(normalizado)

        # Divide o envio ao Supabase em blocos de até 100 registros.
        for inicio in range(0, len(payload), 100):
            lote_payload = payload[inicio: inicio + 100]
            salvos = supabase_upsert(lote_payload)

            # return=representation normalmente devolve os registros salvos.
            # Caso não devolva, contabilizamos o tamanho do payload enviado.
            produtos_salvos += len(salvos) if salvos else len(lote_payload)

        chegou_ultima_pagina = pagina >= total_paginas
        chegou_limite_paginas = (
            max_paginas is not None
            and paginas_processadas >= max_paginas
        )

        if chegou_ultima_pagina or chegou_limite_paginas:
            break

        pagina += 1
        time.sleep(0.6)

    total_lotes = (
        ((indice_global - 1) // tamanho_lote) + 1
        if indice_global
        else 0
    )

    return {
        "status": "ok",
        "filial": filial_normalizada,
        "tiny_utilizado": (
            "TINY_API_KEY_MINAS"
            if filial_normalizada == "mg"
            else "TINY_TOKEN"
        ),
        "produtos_recebidos": produtos_recebidos,
        "produtos_salvos": produtos_salvos,
        "produtos_ignorados_sem_id": produtos_ignorados_sem_id,
        "detalhes_consultados": detalhes_executados,
        "produtos_sem_ncm_nesta_execucao": sem_ncm,
        "produtos_com_ncm_invalido_nesta_execucao": ncm_invalido,
        "total_lotes_nesta_execucao": total_lotes,
        "tamanho_lote": tamanho_lote,
        "paginas_processadas": paginas_processadas,
        "ultima_pagina_lida": pagina,
        "erros_detalhe_amostra": erros_detalhe[:10],
        "observacao": (
            "A sincronização salva somente colunas existentes em "
            "produtos_tributarios. O status_tributario é calculado "
            "virtualmente nas respostas e não é gravado no banco."
        ),
    }


# ============================================================
# IMPORTAÇÃO DE PLANILHAS DO TINY
# ============================================================

_CABECALHOS_IMPORTADOR = {
    "tiny_produto_id": {
        "id", "id produto", "id do produto", "idproduto", "codigo tiny",
        "codigo do tiny", "id tiny", "produto id", "tiny id",
    },
    "codigo": {
        "codigo", "codigo produto", "codigo do produto", "cod produto",
        "cod", "referencia", "ref",
    },
    "sku": {"sku", "codigo sku", "cod sku"},
    "descricao": {
        "descricao", "descricao produto", "descricao do produto", "produto",
        "nome", "nome produto", "nome do produto",
    },
    "ncm": {"ncm", "codigo ncm", "classificacao fiscal"},
    "cest": {"cest", "codigo cest"},
    "fornecedor": {
        "fornecedor", "nome fornecedor", "fornecedor principal",
    },
    "gtin": {
        "gtin", "ean", "gtin ean", "codigo de barras", "codigo barras",
        "codigo de barra", "cod barras",
    },
    "unidade": {"unidade", "un", "unidade medida", "unidade de medida"},
    "origem_mercadoria": {
        "origem", "origem mercadoria", "origem da mercadoria",
    },
    "preco": {
        "preco", "preco venda", "preco de venda", "valor venda",
        "valor de venda",
    },
    "custo": {
        "custo", "preco custo", "preco de custo", "valor custo",
        "valor de custo", "preco medio custo",
    },
}


def _normalizar_cabecalho(valor: Any) -> str:
    texto = safe_str(valor).lower()
    texto = unicodedata.normalize("NFKD", texto)
    texto = "".join(c for c in texto if not unicodedata.combining(c))
    texto = re.sub(r"[^a-z0-9]+", " ", texto)
    return re.sub(r"\s+", " ", texto).strip()


def _campo_por_cabecalho(cabecalho: Any) -> Optional[str]:
    normalizado = _normalizar_cabecalho(cabecalho)
    for campo, aliases in _CABECALHOS_IMPORTADOR.items():
        if normalizado in aliases:
            return campo
    return None


def _normalizar_identificador(valor: Any) -> str:
    texto = safe_str(valor)
    if texto.endswith(".0") and texto[:-2].isdigit():
        texto = texto[:-2]
    return texto.strip()


def _normalizar_gtin(valor: Any) -> Optional[str]:
    texto = _normalizar_identificador(valor)
    digitos = "".join(c for c in texto if c.isdigit())
    return digitos or None


def _normalizar_cest(valor: Any) -> Optional[str]:
    digitos = "".join(c for c in safe_str(valor) if c.isdigit())
    return digitos or None


def _valor_nao_vazio(valor: Any) -> bool:
    if valor is None:
        return False
    if isinstance(valor, str):
        return bool(valor.strip())
    return True


def _detectar_delimitador(texto: str) -> str:
    amostra = texto[:8192]
    try:
        return csv.Sniffer().sniff(amostra, delimiters=";,\t|").delimiter
    except csv.Error:
        return ";" if amostra.count(";") >= amostra.count(",") else ","


def _ler_csv_bytes(conteudo: bytes) -> List[List[Any]]:
    texto = None
    for encoding in ("utf-8-sig", "cp1252", "latin-1"):
        try:
            texto = conteudo.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    if texto is None:
        raise ValueError("Não foi possível identificar a codificação do CSV.")
    delimitador = _detectar_delimitador(texto)
    return [list(linha) for linha in csv.reader(io.StringIO(texto), delimiter=delimitador)]


def _ler_xlsx_bytes(conteudo: bytes) -> List[List[Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ValueError("A dependência openpyxl não está instalada.") from exc
    workbook = load_workbook(io.BytesIO(conteudo), read_only=True, data_only=True)
    planilha = workbook.active
    linhas = [list(linha) for linha in planilha.iter_rows(values_only=True)]
    workbook.close()
    return linhas


def _ler_xls_bytes(conteudo: bytes) -> List[List[Any]]:
    try:
        import xlrd
    except ImportError as exc:
        raise ValueError("A dependência xlrd não está instalada.") from exc
    workbook = xlrd.open_workbook(file_contents=conteudo)
    planilha = workbook.sheet_by_index(0)
    return [planilha.row_values(indice) for indice in range(planilha.nrows)]


def _ler_planilha(nome: str, conteudo: bytes) -> List[List[Any]]:
    extensao = os.path.splitext(nome.lower())[1]
    if extensao == ".csv":
        return _ler_csv_bytes(conteudo)
    if extensao == ".xlsx":
        return _ler_xlsx_bytes(conteudo)
    if extensao == ".xls":
        return _ler_xls_bytes(conteudo)
    raise ValueError("Formato não suportado. Envie arquivo XLS, XLSX ou CSV.")


def _localizar_linha_cabecalho(linhas: List[List[Any]]) -> Tuple[int, Dict[int, str]]:
    melhor_indice = -1
    melhor_mapa: Dict[int, str] = {}
    for indice, linha in enumerate(linhas[:30]):
        mapa: Dict[int, str] = {}
        for coluna, valor in enumerate(linha):
            campo = _campo_por_cabecalho(valor)
            if campo and campo not in mapa.values():
                mapa[coluna] = campo
        if len(mapa) > len(melhor_mapa):
            melhor_indice = indice
            melhor_mapa = mapa
    identificadores = {"tiny_produto_id", "codigo", "sku"}
    if melhor_indice < 0 or not identificadores.intersection(melhor_mapa.values()):
        raise ValueError(
            "Cabeçalho não reconhecido. A planilha precisa ter ID do Tiny, Código ou SKU."
        )
    return melhor_indice, melhor_mapa


def _extrair_registros_planilha(nome: str, conteudo: bytes) -> Tuple[List[Dict[str, Any]], int]:
    linhas = _ler_planilha(nome, conteudo)
    if not linhas:
        return [], 0
    indice_cabecalho, mapa = _localizar_linha_cabecalho(linhas)
    registros: List[Dict[str, Any]] = []
    invalidos = 0
    for numero_linha, linha in enumerate(linhas[indice_cabecalho + 1:], start=indice_cabecalho + 2):
        registro: Dict[str, Any] = {"_arquivo": nome, "_linha": numero_linha}
        for indice_coluna, campo in mapa.items():
            valor = linha[indice_coluna] if indice_coluna < len(linha) else None
            if _valor_nao_vazio(valor):
                registro[campo] = valor
        if not any(_valor_nao_vazio(registro.get(c)) for c in ("tiny_produto_id", "codigo", "sku")):
            if any(_valor_nao_vazio(v) for k, v in registro.items() if not k.startswith("_")):
                invalidos += 1
            continue
        registros.append(registro)
    return registros, invalidos


def _chave_indice(valor: Any) -> str:
    return _normalizar_identificador(valor).casefold()


def importar_produtos_planilhas_tiny(
    filial: str,
    arquivos: List[Dict[str, Any]],
    importar_precos: bool = False,
    atualizar_descricao: bool = False,
) -> Dict[str, Any]:
    filial_normalizada = normalizar_filial(filial)
    if filial_normalizada == "all":
        raise HTTPException(status_code=400, detail="Use filial=sp ou filial=mg.")
    if not arquivos:
        raise HTTPException(status_code=400, detail="Envie pelo menos uma planilha.")

    registros_planilha: List[Dict[str, Any]] = []
    arquivos_processados: List[Dict[str, Any]] = []
    invalidos = 0
    erros_arquivos: List[Dict[str, str]] = []

    for arquivo in arquivos:
        nome = safe_str(arquivo.get("nome")) or "arquivo_sem_nome"
        conteudo = arquivo.get("conteudo") or b""
        if not conteudo:
            erros_arquivos.append({"arquivo": nome, "erro": "Arquivo vazio."})
            continue
        try:
            extraidos, invalidos_arquivo = _extrair_registros_planilha(nome, conteudo)
            registros_planilha.extend(extraidos)
            invalidos += invalidos_arquivo
            arquivos_processados.append({"arquivo": nome, "linhas_validas": len(extraidos)})
        except Exception as exc:
            erros_arquivos.append({"arquivo": nome, "erro": safe_str(exc)})

    if not registros_planilha:
        raise HTTPException(
            status_code=400,
            detail={
                "erro": "Nenhuma linha válida foi encontrada nas planilhas.",
                "arquivos_com_erro": erros_arquivos,
            },
        )

    produtos = supabase_get_todos([
        ("select", "*"),
        ("filial", f"eq.{filial_normalizada}"),
        ("order", "id.asc"),
    ])

    por_tiny = {_chave_indice(p.get("tiny_produto_id")): p for p in produtos if _chave_indice(p.get("tiny_produto_id"))}
    por_sku = {_chave_indice(p.get("sku")): p for p in produtos if _chave_indice(p.get("sku"))}
    por_codigo = {_chave_indice(p.get("codigo")): p for p in produtos if _chave_indice(p.get("codigo"))}

    atualizacoes_por_id: Dict[str, Dict[str, Any]] = {}
    linhas_por_id: Dict[str, int] = {}
    nao_encontrados: List[Dict[str, Any]] = []
    duplicados = 0

    for linha in registros_planilha:
        produto = None
        tiny_id = _chave_indice(linha.get("tiny_produto_id"))
        sku = _chave_indice(linha.get("sku"))
        codigo = _chave_indice(linha.get("codigo"))
        if tiny_id:
            produto = por_tiny.get(tiny_id)
        if produto is None and sku:
            produto = por_sku.get(sku)
        if produto is None and codigo:
            produto = por_codigo.get(codigo)
        if produto is None:
            nao_encontrados.append({
                "arquivo": linha.get("_arquivo"),
                "linha": linha.get("_linha"),
                "tiny_produto_id": _normalizar_identificador(linha.get("tiny_produto_id")),
                "sku": _normalizar_identificador(linha.get("sku")),
                "codigo": _normalizar_identificador(linha.get("codigo")),
                "descricao": safe_str(linha.get("descricao")),
            })
            continue

        produto_id = safe_str(produto.get("id"))
        if produto_id in atualizacoes_por_id:
            duplicados += 1
        payload = atualizacoes_por_id.setdefault(produto_id, dict(produto))
        linhas_por_id[produto_id] = linhas_por_id.get(produto_id, 0) + 1

        if _valor_nao_vazio(linha.get("ncm")):
            ncm = ncm_limpo(linha.get("ncm"))
            if len(ncm) == 8:
                payload["ncm"] = ncm
                payload["status_ia"] = status_ia_inicial(ncm)
        if _valor_nao_vazio(linha.get("cest")):
            payload["cest"] = _normalizar_cest(linha.get("cest"))
        if _valor_nao_vazio(linha.get("fornecedor")):
            payload["fornecedor"] = safe_str(linha.get("fornecedor"))
        if _valor_nao_vazio(linha.get("gtin")):
            payload["gtin"] = _normalizar_gtin(linha.get("gtin"))
        if _valor_nao_vazio(linha.get("unidade")):
            payload["unidade"] = safe_str(linha.get("unidade")).upper()
        if _valor_nao_vazio(linha.get("origem_mercadoria")):
            payload["origem_mercadoria"] = _normalizar_identificador(linha.get("origem_mercadoria"))
        if atualizar_descricao and _valor_nao_vazio(linha.get("descricao")):
            payload["descricao"] = safe_str(linha.get("descricao"))
        if importar_precos:
            if _valor_nao_vazio(linha.get("preco")):
                payload["preco"] = dinheiro_para_float(linha.get("preco"))
            if _valor_nao_vazio(linha.get("custo")):
                payload["custo"] = dinheiro_para_float(linha.get("custo"))
        payload["updated_at"] = agora_iso()
        payload["ultima_sincronizacao"] = agora_iso()

    atualizacoes = list(atualizacoes_por_id.values())
    atualizados = 0
    if atualizacoes:
        try:
            for inicio in range(0, len(atualizacoes), 200):
                bloco = atualizacoes[inicio:inicio + 200]
                salvos = supabase_upsert(bloco)
                atualizados += len(salvos) if salvos else len(bloco)
        except HTTPException as exc:
            detalhe = exc.detail
            if "gtin" in safe_str(detalhe).lower():
                raise HTTPException(
                    status_code=500,
                    detail={
                        "erro": "A coluna gtin ainda não existe em produtos_tributarios.",
                        "acao": "Execute o SQL 01_adicionar_gtin_produtos_tributarios.sql no Supabase e tente novamente.",
                        "detalhe_original": detalhe,
                    },
                )
            raise

    return {
        "status": "ok",
        "filial": filial_normalizada,
        "arquivos_recebidos": len(arquivos),
        "arquivos_processados": arquivos_processados,
        "arquivos_com_erro": erros_arquivos,
        "linhas_lidas": len(registros_planilha) + invalidos,
        "linhas_validas": len(registros_planilha),
        "produtos_atualizados": atualizados,
        "duplicados": duplicados,
        "linhas_invalidas": invalidos,
        "nao_encontrados": len(nao_encontrados),
        "nao_encontrados_amostra": nao_encontrados[:100],
        "importou_precos": importar_precos,
        "atualizou_descricao": atualizar_descricao,
    }


# ============================================================
# ENRIQUECIMENTO INCREMENTAL DOS PRODUTOS
# ============================================================

def _texto_erro_tiny(exc: Exception) -> str:
    """Transforma erros do Tiny em texto para identificação de bloqueio."""
    if isinstance(exc, HTTPException):
        return safe_str(exc.detail).lower()
    return safe_str(exc).lower()


def _tiny_bloqueado(texto_erro: str) -> bool:
    """Identifica as mensagens mais comuns de limite de acesso do Tiny."""
    texto = safe_str(texto_erro).lower()
    return any(
        trecho in texto
        for trecho in (
            "api bloqueada",
            "excedido o número de acessos",
            "excedido o numero de acessos",
            "número de acessos a api",
            "numero de acessos a api",
        )
    )


def _atualizar_produto_enriquecido(
    produto_id: str,
    filial: str,
    dados: Dict[str, Any],
) -> List[Dict[str, Any]]:
    """
    Atualiza um produto enriquecido por ID e filial.

    Diferente do PATCH manual, esta função pode atualizar também
    ultima_sincronizacao, pois esse campo faz parte do processo automático.
    """
    produto_id_normalizado = safe_str(produto_id)
    filial_normalizada = normalizar_filial(filial)

    if not produto_id_normalizado:
        raise HTTPException(
            status_code=400,
            detail="produto_id não informado para o enriquecimento.",
        )

    payload = limpar_payload_banco(dados)
    payload["updated_at"] = agora_iso()

    response = requests.patch(
        f"{SUPABASE_URL}/rest/v1/{TABELA_PRODUTOS}",
        headers=supabase_headers("return=representation"),
        params={
            "id": f"eq.{produto_id_normalizado}",
            "filial": f"eq.{filial_normalizada}",
        },
        json=payload,
        timeout=90,
    )

    if response.status_code >= 400:
        raise HTTPException(
            status_code=500,
            detail={
                "erro": "Erro ao atualizar produto enriquecido.",
                "status_code": response.status_code,
                "resposta": response.text,
            },
        )

    try:
        registros = response.json()
    except ValueError:
        registros = []

    return registros if isinstance(registros, list) else []


def _listar_produtos_pendentes_enriquecimento(
    filial: str,
) -> Tuple[List[Dict[str, Any]], int, int]:
    """
    Lista produtos sem NCM que ainda podem ser consultados no Tiny.

    Produtos já consultados e que realmente não possuem NCM no Tiny são
    marcados como REVISAO_MANUAL e não entram novamente na próxima rodada.
    """
    filial_normalizada = normalizar_filial(filial)

    registros = supabase_get_todos(
        [
            (
                "select",
                (
                    "id,filial,tiny_produto_id,codigo,sku,descricao,ncm,cest,"
                    "origem_mercadoria,unidade,preco,custo,fornecedor,ativo,"
                    "status_ia,percentual_confianca,justificativa_ia,"
                    "data_analise,lote,ultima_sincronizacao"
                ),
            ),
            ("filial", f"eq.{filial_normalizada}"),
            ("order", "ultima_sincronizacao.asc.nullsfirst,id.asc"),
        ],
        tamanho_pagina=1000,
    )

    sem_ncm = [
        registro
        for registro in registros
        if not ncm_limpo(registro.get("ncm"))
    ]

    ja_consultados_sem_ncm = [
        registro
        for registro in sem_ncm
        if safe_str(registro.get("status_ia")).upper() == "REVISAO_MANUAL"
        and "tiny não informou ncm" in safe_str(
            registro.get("justificativa_ia")
        ).lower()
    ]

    pendentes = [
        registro
        for registro in sem_ncm
        if registro not in ja_consultados_sem_ncm
        and bool(safe_str(registro.get("tiny_produto_id")))
    ]

    return pendentes, len(sem_ncm), len(ja_consultados_sem_ncm)


def enriquecer_produtos(
    filial: str,
    limite: int = 50,
    intervalo_segundos: float = 0.7,
) -> Dict[str, Any]:
    """
    Busca no Tiny os detalhes dos produtos ainda sem NCM.

    O processo é incremental e seguro para a API:
    - trabalha somente com sp ou mg;
    - consulta no máximo o limite informado;
    - espera entre as chamadas;
    - interrompe imediatamente se o Tiny bloquear a API;
    - não consulta novamente produtos que já foram verificados e continuam
      sem NCM no cadastro do Tiny.
    """
    filial_normalizada = normalizar_filial(filial)

    if filial_normalizada == "all":
        raise HTTPException(
            status_code=400,
            detail="Use filial=sp ou filial=mg.",
        )

    if limite < 1 or limite > 200:
        raise HTTPException(
            status_code=400,
            detail="limite deve estar entre 1 e 200.",
        )

    if intervalo_segundos < 0.5 or intervalo_segundos > 10:
        raise HTTPException(
            status_code=400,
            detail="intervalo_segundos deve estar entre 0.5 e 10.",
        )

    pendentes, total_sem_ncm_antes, ja_consultados_sem_ncm = (
        _listar_produtos_pendentes_enriquecimento(filial_normalizada)
    )
    selecionados = pendentes[:limite]

    if not selecionados:
        return {
            "status": "ok",
            "filial": filial_normalizada,
            "tiny_utilizado": (
                "TINY_API_KEY_MINAS"
                if filial_normalizada == "mg"
                else "TINY_TOKEN"
            ),
            "total_sem_ncm_antes": total_sem_ncm_antes,
            "pendentes_para_consulta_antes": len(pendentes),
            "ja_consultados_sem_ncm": ja_consultados_sem_ncm,
            "produtos_selecionados": 0,
            "produtos_consultados": 0,
            "produtos_atualizados": 0,
            "produtos_com_ncm_encontrado": 0,
            "produtos_sem_ncm_no_tiny": 0,
            "erros_quantidade": 0,
            "erros_amostra": [],
            "api_tiny_bloqueada": False,
            "total_sem_ncm_depois": total_sem_ncm_antes,
            "pendentes_para_consulta_depois": len(pendentes),
            "concluido": len(pendentes) == 0,
            "mensagem": "Nenhum produto pendente para consultar no Tiny.",
        }

    consultados = 0
    atualizados = 0
    com_ncm_encontrado = 0
    sem_ncm_no_tiny = 0
    api_bloqueada = False
    erros: List[Dict[str, Any]] = []

    for indice, produto_banco in enumerate(selecionados):
        produto_id = safe_str(produto_banco.get("id"))
        tiny_produto_id = safe_str(produto_banco.get("tiny_produto_id"))

        if not produto_id or not tiny_produto_id:
            erros.append(
                {
                    "produto_id": produto_id or None,
                    "tiny_produto_id": tiny_produto_id or None,
                    "descricao": produto_banco.get("descricao"),
                    "erro": "Produto sem ID interno ou tiny_produto_id.",
                }
            )
            continue

        try:
            detalhe = obter_produto_tiny(
                tiny_produto_id,
                filial_normalizada,
            )
            consultados += 1

            if not detalhe:
                erros.append(
                    {
                        "produto_id": produto_id,
                        "tiny_produto_id": tiny_produto_id,
                        "descricao": produto_banco.get("descricao"),
                        "erro": "Tiny não retornou detalhes do produto.",
                    }
                )
            else:
                normalizado = normalizar_produto(
                    produto_pesquisa=produto_banco,
                    filial=filial_normalizada,
                    lote=int(produto_banco.get("lote") or 1),
                    produto_detalhado=detalhe,
                )

                ncm_encontrado = ncm_limpo(normalizado.get("ncm"))
                agora = agora_iso()

                if len(ncm_encontrado) == 8:
                    normalizado["ncm"] = ncm_encontrado
                    normalizado["status_ia"] = "PENDENTE"
                    normalizado["percentual_confianca"] = None
                    normalizado["justificativa_ia"] = None
                    normalizado["data_analise"] = None
                    com_ncm_encontrado += 1
                else:
                    normalizado["ncm"] = None
                    normalizado["status_ia"] = "REVISAO_MANUAL"
                    normalizado["percentual_confianca"] = 0
                    normalizado["justificativa_ia"] = (
                        "Tiny não informou NCM no detalhe do produto; "
                        "necessária revisão manual do cadastro."
                    )
                    normalizado["data_analise"] = agora
                    sem_ncm_no_tiny += 1

                normalizado["ultima_sincronizacao"] = agora
                normalizado["updated_at"] = agora

                registros_atualizados = _atualizar_produto_enriquecido(
                    produto_id=produto_id,
                    filial=filial_normalizada,
                    dados=normalizado,
                )
                atualizados += (
                    len(registros_atualizados)
                    if registros_atualizados
                    else 1
                )

        except HTTPException as exc:
            texto_erro = _texto_erro_tiny(exc)
            erros.append(
                {
                    "produto_id": produto_id,
                    "tiny_produto_id": tiny_produto_id,
                    "descricao": produto_banco.get("descricao"),
                    "erro": safe_str(exc.detail)[:500],
                }
            )

            if _tiny_bloqueado(texto_erro):
                api_bloqueada = True
                break

        except Exception as exc:
            texto_erro = _texto_erro_tiny(exc)
            erros.append(
                {
                    "produto_id": produto_id,
                    "tiny_produto_id": tiny_produto_id,
                    "descricao": produto_banco.get("descricao"),
                    "erro": safe_str(exc)[:500],
                }
            )

            if _tiny_bloqueado(texto_erro):
                api_bloqueada = True
                break

        if indice < len(selecionados) - 1:
            time.sleep(intervalo_segundos)

    pendentes_depois, total_sem_ncm_depois, ja_consultados_depois = (
        _listar_produtos_pendentes_enriquecimento(filial_normalizada)
    )

    status_execucao = "ok"
    if api_bloqueada:
        status_execucao = "bloqueado"
    elif erros:
        status_execucao = "parcial"

    return {
        "status": status_execucao,
        "filial": filial_normalizada,
        "tiny_utilizado": (
            "TINY_API_KEY_MINAS"
            if filial_normalizada == "mg"
            else "TINY_TOKEN"
        ),
        "total_sem_ncm_antes": total_sem_ncm_antes,
        "pendentes_para_consulta_antes": len(pendentes),
        "ja_consultados_sem_ncm_antes": ja_consultados_sem_ncm,
        "produtos_selecionados": len(selecionados),
        "produtos_consultados": consultados,
        "produtos_atualizados": atualizados,
        "produtos_com_ncm_encontrado": com_ncm_encontrado,
        "produtos_sem_ncm_no_tiny": sem_ncm_no_tiny,
        "erros_quantidade": len(erros),
        "erros_amostra": erros[:10],
        "api_tiny_bloqueada": api_bloqueada,
        "total_sem_ncm_depois": total_sem_ncm_depois,
        "pendentes_para_consulta_depois": len(pendentes_depois),
        "ja_consultados_sem_ncm_depois": ja_consultados_depois,
        "concluido": len(pendentes_depois) == 0,
        "observacao": (
            "Execute novamente para continuar pelos próximos produtos. "
            "Produtos já consultados e sem NCM no Tiny ficam marcados para "
            "revisão manual e não são consultados novamente."
        ),
    }


# ============================================================
# CONSULTAS
# ============================================================

def montar_filtro_busca(busca: str) -> str:
    """
    Prepara um termo simples para o filtro OR do PostgREST.
    """
    termo = safe_str(busca)
    termo = termo.replace(",", " ").replace("(", " ").replace(")", " ")
    termo = " ".join(termo.split())
    return termo


def listar_produtos(
    filial: str,
    status_tributario: Optional[str],
    status_ia: Optional[str],
    lote: Optional[int],
    busca: Optional[str],
    limite: int,
    offset: int,
) -> List[Dict[str, Any]]:
    if limite < 1:
        raise HTTPException(
            status_code=400,
            detail="limite deve ser maior ou igual a 1.",
        )

    if limite > 10000:
        limite = 10000

    if offset < 0:
        raise HTTPException(
            status_code=400,
            detail="offset não pode ser negativo.",
        )

    filial_normalizada = normalizar_filial(filial)

    # status_tributario é virtual. Quando esse filtro é usado,
    # buscamos uma faixa maior e filtramos em memória.
    usa_filtro_virtual = bool(status_tributario)

    limite_consulta = 10000 if usa_filtro_virtual else limite
    offset_consulta = 0 if usa_filtro_virtual else offset

    params: List[Tuple[str, str]] = [
        ("select", "*"),
        ("order", "lote.asc,descricao.asc"),
        ("limit", str(limite_consulta)),
        ("offset", str(offset_consulta)),
    ]

    if filial_normalizada != "all":
        params.append(("filial", f"eq.{filial_normalizada}"))

    if status_ia:
        params.append(("status_ia", f"eq.{safe_str(status_ia)}"))

    if lote is not None:
        params.append(("lote", f"eq.{lote}"))

    if busca:
        termo = montar_filtro_busca(busca)

        if termo:
            params.append(
                (
                    "or",
                    (
                        f"(descricao.ilike.*{termo}*,"
                        f"codigo.ilike.*{termo}*,"
                        f"sku.ilike.*{termo}*,"
                        f"ncm.ilike.*{termo}*)"
                    ),
                )
            )

    registros = [
        adicionar_status_tributario_virtual(registro)
        for registro in supabase_get(params)
    ]

    if status_tributario:
        status_desejado = safe_str(status_tributario).lower()

        registros = [
            registro
            for registro in registros
            if safe_str(registro.get("status_tributario")).lower()
            == status_desejado
        ]

        registros = registros[offset: offset + limite]

    return registros


def obter_resumo(filial: str) -> Dict[str, Any]:
    filial_normalizada = normalizar_filial(filial)

    params: List[Tuple[str, str]] = [
        (
            "select",
            (
                "id,ncm,lote,status_ia,percentual_confianca,"
                "data_analise,ativo"
            ),
        ),
        ("limit", "10000"),
    ]

    if filial_normalizada != "all":
        params.append(("filial", f"eq.{filial_normalizada}"))

    registros = supabase_get(params)

    por_status_tributario: Dict[str, int] = {}
    por_status_ia: Dict[str, int] = {}
    lotes = set()

    com_ncm = 0
    sem_ncm = 0
    ncm_invalido = 0
    ativos = 0
    analisados = 0
    confiancas: List[float] = []

    for registro in registros:
        status_tributario_calculado = status_ncm(registro.get("ncm"))
        por_status_tributario[status_tributario_calculado] = (
            por_status_tributario.get(status_tributario_calculado, 0) + 1
        )

        if status_tributario_calculado == "sem_ncm":
            sem_ncm += 1
        elif status_tributario_calculado == "ncm_invalido":
            ncm_invalido += 1
        else:
            com_ncm += 1

        status_ia_valor = (
            safe_str(registro.get("status_ia")).upper()
            or "SEM_STATUS"
        )

        por_status_ia[status_ia_valor] = (
            por_status_ia.get(status_ia_valor, 0) + 1
        )

        if registro.get("lote") is not None:
            lotes.add(registro.get("lote"))

        if registro.get("ativo") is True:
            ativos += 1

        if registro.get("data_analise"):
            analisados += 1

        confianca = registro.get("percentual_confianca")

        if confianca is not None:
            try:
                confiancas.append(float(confianca))
            except (TypeError, ValueError):
                pass

    total = len(registros)

    confianca_media = (
        round(sum(confiancas) / len(confiancas), 2)
        if confiancas
        else 0.0
    )

    return {
        "status": "ok",
        "filial": filial_normalizada,
        "total_produtos": total,
        "produtos_ativos": ativos,
        "produtos_com_ncm": com_ncm,
        "produtos_sem_ncm": sem_ncm,
        "produtos_com_ncm_invalido": ncm_invalido,
        "produtos_analisados": analisados,
        "percentual_confianca_medio": confianca_media,
        "total_lotes": len(lotes),
        "por_status_tributario": por_status_tributario,
        "por_status_ia": por_status_ia,
    }


def listar_lotes(filial: str) -> List[Dict[str, Any]]:
    registros = listar_produtos(
        filial=filial,
        status_tributario=None,
        status_ia=None,
        lote=None,
        busca=None,
        limite=10000,
        offset=0,
    )

    agrupado: Dict[int, Dict[str, Any]] = {}

    for registro in registros:
        try:
            numero = int(registro.get("lote") or 0)
        except (TypeError, ValueError):
            numero = 0

        if numero <= 0:
            continue

        item = agrupado.setdefault(
            numero,
            {
                "lote": numero,
                "total_produtos": 0,
                "pendentes": 0,
                "processando": 0,
                "concluidos": 0,
                "erros": 0,
                "revisao_manual": 0,
                "sem_ncm": 0,
                "ncm_invalido": 0,
                "analisados": 0,
            },
        )

        item["total_produtos"] += 1

        status_ia_valor = safe_str(
            registro.get("status_ia")
        ).upper()

        status_tributario_valor = safe_str(
            registro.get("status_tributario")
        ).lower()

        if status_ia_valor == "PENDENTE":
            item["pendentes"] += 1
        elif status_ia_valor == "PROCESSANDO":
            item["processando"] += 1
        elif status_ia_valor in {"CONCLUIDO", "CONCLUÍDO"}:
            item["concluidos"] += 1
        elif status_ia_valor == "ERRO":
            item["erros"] += 1
        elif status_ia_valor == "REVISAO_MANUAL":
            item["revisao_manual"] += 1

        if status_tributario_valor == "sem_ncm":
            item["sem_ncm"] += 1
        elif status_tributario_valor == "ncm_invalido":
            item["ncm_invalido"] += 1

        if registro.get("data_analise"):
            item["analisados"] += 1

    resultado = list(agrupado.values())

    for item in resultado:
        if item["processando"] > 0:
            item["status"] = "PROCESSANDO"
        elif item["pendentes"] > 0:
            item["status"] = "PENDENTE"
        elif (
            item["erros"] > 0
            or item["revisao_manual"] > 0
            or item["ncm_invalido"] > 0
        ):
            item["status"] = "REVISAO"
        else:
            item["status"] = "CONCLUIDO"

    return sorted(
        resultado,
        key=lambda item: item["lote"],
    )


def obter_lote(
    filial: str,
    numero_lote: int,
) -> List[Dict[str, Any]]:
    """
    Função auxiliar para o endpoint GET /tributario/lote/{numero}.
    """
    return listar_produtos(
        filial=filial,
        status_tributario=None,
        status_ia=None,
        lote=numero_lote,
        busca=None,
        limite=10000,
        offset=0,
    )


def atualizar_produto(
    produto_id: str,
    dados: Dict[str, Any],
) -> List[Dict[str, Any]]:
    """
    Função auxiliar para o endpoint PATCH /tributario/produto/{produto_id}.
    """
    return supabase_patch(produto_id, dados)

# ============================================================
# REGRAS TRIBUTÁRIAS POR NCM
# ============================================================

TABELA_REGRAS = "tributacao_ncm"

COLUNAS_TRIBUTACAO_NCM = {
    "id", "ncm", "descricao_ncm", "cest", "ex_tipi",
    "regime_tributario", "uf_origem", "uf_destino", "tipo_operacao",
    "finalidade", "consumidor_final", "contribuinte_icms", "cfop",
    "cst_icms", "csosn", "aliquota_icms", "reducao_base_icms",
    "possui_icms_st", "mva", "aliquota_icms_st", "reducao_base_icms_st",
    "possui_fcp", "aliquota_fcp", "aliquota_fcp_st", "possui_difal",
    "aliquota_interna_destino", "aliquota_interestadual", "cst_ipi",
    "aliquota_ipi", "codigo_enquadramento_ipi", "cst_pis", "aliquota_pis",
    "cst_cofins", "aliquota_cofins", "pis_cofins_monofasico",
    "substituicao_tributaria", "beneficio_fiscal", "codigo_beneficio_fiscal",
    "desoneracao_icms", "status", "percentual_confianca", "justificativa_ia",
    "observacoes", "fonte", "link_fonte", "data_consulta_fonte",
    "vigencia_inicio", "vigencia_fim", "ativo", "revisao_manual",
    "created_at", "updated_at",
}

COLUNAS_REGRAS_EDITAVEIS = COLUNAS_TRIBUTACAO_NCM - {"id", "created_at", "updated_at"}


def _supabase_request_tabela(
    method: str,
    tabela: str,
    *,
    params: Optional[Any] = None,
    json: Optional[Any] = None,
    prefer: Optional[str] = None,
    timeout: int = 90,
) -> Any:
    response = requests.request(
        method=method,
        url=f"{SUPABASE_URL}/rest/v1/{tabela}",
        headers=supabase_headers(prefer),
        params=params,
        json=json,
        timeout=timeout,
    )

    if response.status_code >= 400:
        raise HTTPException(
            status_code=500,
            detail={
                "erro": f"Erro ao acessar a tabela {tabela}.",
                "metodo": method.upper(),
                "status_code": response.status_code,
                "resposta": response.text,
            },
        )

    if response.status_code == 204 or not response.text.strip():
        return []

    try:
        return response.json()
    except ValueError:
        return response.text


def _normalizar_uf(valor: Any, campo: str) -> str:
    uf = safe_str(valor).upper()
    if len(uf) != 2 or not uf.isalpha():
        raise HTTPException(status_code=400, detail=f"{campo} deve possuir 2 letras.")
    return uf


def _normalizar_ncm_obrigatorio(valor: Any) -> str:
    ncm = ncm_limpo(valor)
    if len(ncm) != 8:
        raise HTTPException(status_code=400, detail="ncm deve possuir exatamente 8 números.")
    return ncm


def _limpar_regra_payload(dados: Dict[str, Any], parcial: bool = False) -> Dict[str, Any]:
    payload = {
        chave: valor
        for chave, valor in dict(dados).items()
        if chave in COLUNAS_REGRAS_EDITAVEIS and valor is not None
    }

    if not parcial or "ncm" in payload:
        if "ncm" not in payload:
            raise HTTPException(status_code=400, detail="O campo ncm é obrigatório.")
        payload["ncm"] = _normalizar_ncm_obrigatorio(payload["ncm"])

    for campo in ("uf_origem", "uf_destino"):
        if campo in payload:
            payload[campo] = _normalizar_uf(payload[campo], campo)

    for campo in (
        "regime_tributario", "tipo_operacao", "finalidade", "status",
    ):
        if campo in payload:
            payload[campo] = safe_str(payload[campo]).lower()

    if "cest" in payload:
        payload["cest"] = ncm_limpo(payload["cest"]) or None

    if "percentual_confianca" in payload:
        try:
            confianca = float(payload["percentual_confianca"])
        except (TypeError, ValueError):
            raise HTTPException(status_code=400, detail="percentual_confianca inválido.")
        if confianca < 0 or confianca > 100:
            raise HTTPException(status_code=400, detail="percentual_confianca deve estar entre 0 e 100.")
        payload["percentual_confianca"] = confianca

    payload["updated_at"] = agora_iso()
    return payload


def listar_regras(
    ncm: Optional[str] = None,
    regime_tributario: Optional[str] = None,
    uf_origem: Optional[str] = None,
    uf_destino: Optional[str] = None,
    tipo_operacao: Optional[str] = None,
    finalidade: Optional[str] = None,
    status: Optional[str] = None,
    ativo: Optional[bool] = True,
    limite: int = 100,
    offset: int = 0,
) -> List[Dict[str, Any]]:
    if limite < 1 or limite > 10000:
        raise HTTPException(status_code=400, detail="limite deve estar entre 1 e 10000.")
    if offset < 0:
        raise HTTPException(status_code=400, detail="offset não pode ser negativo.")

    params: List[Tuple[str, str]] = [
        ("select", "*"),
        ("order", "ncm.asc,uf_origem.asc,uf_destino.asc,vigencia_inicio.desc"),
        ("limit", str(limite)),
        ("offset", str(offset)),
    ]

    if ncm:
        params.append(("ncm", f"eq.{_normalizar_ncm_obrigatorio(ncm)}"))
    if regime_tributario:
        params.append(("regime_tributario", f"eq.{safe_str(regime_tributario).lower()}"))
    if uf_origem:
        params.append(("uf_origem", f"eq.{_normalizar_uf(uf_origem, 'uf_origem')}"))
    if uf_destino:
        params.append(("uf_destino", f"eq.{_normalizar_uf(uf_destino, 'uf_destino')}"))
    if tipo_operacao:
        params.append(("tipo_operacao", f"eq.{safe_str(tipo_operacao).lower()}"))
    if finalidade:
        params.append(("finalidade", f"eq.{safe_str(finalidade).lower()}"))
    if status:
        params.append(("status", f"eq.{safe_str(status).lower()}"))
    if ativo is not None:
        params.append(("ativo", f"eq.{str(bool(ativo)).lower()}"))

    dados = _supabase_request_tabela("GET", TABELA_REGRAS, params=params)
    return dados if isinstance(dados, list) else []


def obter_regra(regra_id: str) -> Dict[str, Any]:
    dados = _supabase_request_tabela(
        "GET",
        TABELA_REGRAS,
        params={"select": "*", "id": f"eq.{regra_id}", "limit": "1"},
    )
    if not dados:
        raise HTTPException(status_code=404, detail="Regra tributária não encontrada.")
    return dados[0]


def criar_regra(dados: Dict[str, Any]) -> Dict[str, Any]:
    payload = _limpar_regra_payload(dados, parcial=False)
    payload.setdefault("regime_tributario", "lucro_presumido")
    payload.setdefault("uf_origem", "SP")
    payload.setdefault("uf_destino", "SP")
    payload.setdefault("tipo_operacao", "venda")
    payload.setdefault("finalidade", "revenda")
    payload.setdefault("consumidor_final", False)
    payload.setdefault("contribuinte_icms", True)
    payload.setdefault("status", "pendente")
    payload.setdefault("ativo", True)
    payload.setdefault("revisao_manual", False)

    dados_salvos = _supabase_request_tabela(
        "POST",
        TABELA_REGRAS,
        json=payload,
        prefer="return=representation",
        timeout=120,
    )
    if not dados_salvos:
        raise HTTPException(status_code=500, detail="A regra foi enviada, mas o Supabase não retornou o registro.")
    return dados_salvos[0]


def atualizar_regra(regra_id: str, dados: Dict[str, Any]) -> Dict[str, Any]:
    obter_regra(regra_id)
    payload = _limpar_regra_payload(dados, parcial=True)
    if not payload or set(payload.keys()) == {"updated_at"}:
        raise HTTPException(status_code=400, detail="Nenhum campo válido foi enviado.")

    atualizados = _supabase_request_tabela(
        "PATCH",
        TABELA_REGRAS,
        params={"id": f"eq.{regra_id}"},
        json=payload,
        prefer="return=representation",
    )
    if not atualizados:
        raise HTTPException(status_code=404, detail="Regra tributária não encontrada.")
    return atualizados[0]


def excluir_regra(regra_id: str) -> Dict[str, Any]:
    regra = obter_regra(regra_id)
    _supabase_request_tabela(
        "DELETE",
        TABELA_REGRAS,
        params={"id": f"eq.{regra_id}"},
        prefer="return=minimal",
    )
    return {"status": "ok", "mensagem": "Regra tributária excluída.", "id": regra["id"]}


def buscar_regra_aplicavel(
    ncm: str,
    regime_tributario: str,
    uf_origem: str,
    uf_destino: str,
    tipo_operacao: str = "venda",
    finalidade: str = "revenda",
    consumidor_final: bool = False,
    contribuinte_icms: bool = True,
    data_referencia: Optional[str] = None,
) -> Dict[str, Any]:
    from datetime import date

    ncm_normalizado = _normalizar_ncm_obrigatorio(ncm)
    uf_origem_normalizada = _normalizar_uf(uf_origem, "uf_origem")
    uf_destino_normalizada = _normalizar_uf(uf_destino, "uf_destino")

    try:
        referencia = date.fromisoformat(data_referencia) if data_referencia else date.today()
    except ValueError:
        raise HTTPException(status_code=400, detail="data_referencia deve estar no formato YYYY-MM-DD.")

    params: List[Tuple[str, str]] = [
        ("select", "*"),
        ("ncm", f"eq.{ncm_normalizado}"),
        ("regime_tributario", f"eq.{safe_str(regime_tributario).lower()}"),
        ("uf_origem", f"eq.{uf_origem_normalizada}"),
        ("uf_destino", f"eq.{uf_destino_normalizada}"),
        ("tipo_operacao", f"eq.{safe_str(tipo_operacao).lower()}"),
        ("finalidade", f"eq.{safe_str(finalidade).lower()}"),
        ("consumidor_final", f"eq.{str(bool(consumidor_final)).lower()}"),
        ("contribuinte_icms", f"eq.{str(bool(contribuinte_icms)).lower()}"),
        ("ativo", "eq.true"),
        ("order", "vigencia_inicio.desc.nullslast,created_at.desc"),
        ("limit", "100"),
    ]

    candidatas = _supabase_request_tabela("GET", TABELA_REGRAS, params=params)

    aplicaveis: List[Dict[str, Any]] = []
    for regra in candidatas or []:
        inicio = regra.get("vigencia_inicio")
        fim = regra.get("vigencia_fim")
        inicio_data = date.fromisoformat(inicio) if inicio else None
        fim_data = date.fromisoformat(fim) if fim else None
        if inicio_data and referencia < inicio_data:
            continue
        if fim_data and referencia > fim_data:
            continue
        aplicaveis.append(regra)

    if not aplicaveis:
        raise HTTPException(
            status_code=404,
            detail={
                "erro": "Nenhuma regra tributária aplicável foi encontrada.",
                "contexto": {
                    "ncm": ncm_normalizado,
                    "regime_tributario": safe_str(regime_tributario).lower(),
                    "uf_origem": uf_origem_normalizada,
                    "uf_destino": uf_destino_normalizada,
                    "tipo_operacao": safe_str(tipo_operacao).lower(),
                    "finalidade": safe_str(finalidade).lower(),
                    "consumidor_final": consumidor_final,
                    "contribuinte_icms": contribuinte_icms,
                    "data_referencia": referencia.isoformat(),
                },
            },
        )

    return {
        "status": "ok",
        "data_referencia": referencia.isoformat(),
        "regra": aplicaveis[0],
        "total_regras_aplicaveis": len(aplicaveis),
    }


# ============================================================
# AUDITORIA TRIBUTÁRIA
# ============================================================

TABELA_AUDITORIA = "auditoria_tributaria"


def _valor_float_seguro(valor: Any) -> float:
    try:
        return float(valor or 0)
    except (TypeError, ValueError):
        return 0.0


def _regra_esta_completa(regra: Dict[str, Any]) -> bool:
    """Valida os campos mínimos para considerar a regra utilizável."""
    possui_icms = bool(safe_str(regra.get("cst_icms")) or safe_str(regra.get("csosn")))
    return all(
        [
            safe_str(regra.get("cfop")),
            possui_icms,
            safe_str(regra.get("cst_pis")),
            safe_str(regra.get("cst_cofins")),
        ]
    )


def _status_regra_para_auditoria(regra: Dict[str, Any]) -> Tuple[str, str]:
    status_regra = safe_str(regra.get("status")).lower()
    revisao_manual = bool(regra.get("revisao_manual"))
    completa = _regra_esta_completa(regra)

    if revisao_manual:
        return "revisar", "A regra tributária exige revisão manual."

    if not completa:
        return "revisar", "A regra foi encontrada, mas possui campos tributários obrigatórios incompletos."

    if status_regra in {"divergente", "erro", "revisao", "revisão", "pendente"}:
        return "revisar", f"A regra tributária está com status '{status_regra or 'pendente'}'."

    return "ok", "Produto vinculado a uma regra tributária aplicável e completa."


def _buscar_regras_contexto_auditoria(
    regime_tributario: str,
    uf_origem: str,
    uf_destino: str,
    tipo_operacao: str,
    finalidade: str,
    consumidor_final: bool,
    contribuinte_icms: bool,
) -> List[Dict[str, Any]]:
    params: List[Tuple[str, str]] = [
        ("select", "*"),
        ("regime_tributario", f"eq.{safe_str(regime_tributario).lower()}"),
        ("uf_origem", f"eq.{_normalizar_uf(uf_origem, 'uf_origem')}"),
        ("uf_destino", f"eq.{_normalizar_uf(uf_destino, 'uf_destino')}"),
        ("tipo_operacao", f"eq.{safe_str(tipo_operacao).lower()}"),
        ("finalidade", f"eq.{safe_str(finalidade).lower()}"),
        ("consumidor_final", f"eq.{str(bool(consumidor_final)).lower()}"),
        ("contribuinte_icms", f"eq.{str(bool(contribuinte_icms)).lower()}"),
        ("ativo", "eq.true"),
        ("order", "ncm.asc,vigencia_inicio.desc.nullslast,created_at.desc"),
        ("limit", "10000"),
    ]
    dados = _supabase_request_tabela("GET", TABELA_REGRAS, params=params)
    return dados if isinstance(dados, list) else []


def _regra_vigente(regra: Dict[str, Any], referencia: date) -> bool:
    try:
        inicio = date.fromisoformat(regra["vigencia_inicio"]) if regra.get("vigencia_inicio") else None
        fim = date.fromisoformat(regra["vigencia_fim"]) if regra.get("vigencia_fim") else None
    except (TypeError, ValueError):
        return False

    if inicio and referencia < inicio:
        return False
    if fim and referencia > fim:
        return False
    return True


def _salvar_auditorias(payload: List[Dict[str, Any]]) -> int:
    if not payload:
        return 0

    total = 0
    for inicio in range(0, len(payload), 200):
        bloco = payload[inicio: inicio + 200]
        salvos = _supabase_request_tabela(
            "POST",
            TABELA_AUDITORIA,
            json=bloco,
            prefer="return=representation",
            timeout=120,
        )
        total += len(salvos) if isinstance(salvos, list) else len(bloco)
    return total


def auditar_tributacao(
    filial: str,
    regime_tributario: str = "lucro_presumido",
    uf_origem: Optional[str] = None,
    uf_destino: Optional[str] = None,
    tipo_operacao: str = "venda",
    finalidade: str = "revenda",
    consumidor_final: bool = False,
    contribuinte_icms: bool = True,
    data_referencia: Optional[str] = None,
    limite: int = 10000,
) -> Dict[str, Any]:
    filial_normalizada = normalizar_filial(filial)
    if filial_normalizada == "all":
        raise HTTPException(status_code=400, detail="Para auditar, use filial=sp ou filial=mg.")

    if limite < 1 or limite > 10000:
        raise HTTPException(status_code=400, detail="limite deve estar entre 1 e 10000.")

    uf_padrao = "MG" if filial_normalizada == "mg" else "SP"
    origem = _normalizar_uf(uf_origem or uf_padrao, "uf_origem")
    destino = _normalizar_uf(uf_destino or uf_padrao, "uf_destino")

    try:
        referencia = date.fromisoformat(data_referencia) if data_referencia else date.today()
    except ValueError:
        raise HTTPException(status_code=400, detail="data_referencia deve estar no formato YYYY-MM-DD.")

    produtos = listar_produtos(
        filial=filial_normalizada,
        status_tributario=None,
        status_ia=None,
        lote=None,
        busca=None,
        limite=limite,
        offset=0,
    )

    regras = _buscar_regras_contexto_auditoria(
        regime_tributario=regime_tributario,
        uf_origem=origem,
        uf_destino=destino,
        tipo_operacao=tipo_operacao,
        finalidade=finalidade,
        consumidor_final=consumidor_final,
        contribuinte_icms=contribuinte_icms,
    )

    regras_por_ncm: Dict[str, List[Dict[str, Any]]] = {}
    for regra in regras:
        ncm_regra = ncm_limpo(regra.get("ncm"))
        if ncm_regra and _regra_vigente(regra, referencia):
            regras_por_ncm.setdefault(ncm_regra, []).append(regra)

    agora = agora_iso()
    auditorias: List[Dict[str, Any]] = []
    produtos_atualizados: List[Dict[str, Any]] = []
    contadores = {"ok": 0, "revisar": 0, "sem_regra": 0, "erro": 0}
    confiancas: List[float] = []

    for produto in produtos:
        produto_id = safe_str(produto.get("id"))
        ncm_produto = ncm_limpo(produto.get("ncm"))
        regra: Optional[Dict[str, Any]] = None
        status = "sem_regra"
        confianca = 0.0
        justificativa = "Produto sem NCM informado."

        if ncm_produto and len(ncm_produto) != 8:
            status = "revisar"
            justificativa = "O NCM do produto é inválido; deve possuir exatamente 8 números."
        elif ncm_produto:
            candidatas = regras_por_ncm.get(ncm_produto, [])
            regra = candidatas[0] if candidatas else None
            if regra:
                status, justificativa = _status_regra_para_auditoria(regra)
                confianca = _valor_float_seguro(regra.get("percentual_confianca"))
            else:
                status = "sem_regra"
                justificativa = "Nenhuma regra tributária aplicável foi encontrada para este NCM e contexto."

        contadores[status] += 1
        if confianca > 0:
            confiancas.append(confianca)

        auditorias.append(
            {
                "produto_id": produto_id,
                "filial": filial_normalizada,
                "regra_id": regra.get("id") if regra else None,
                "tiny_produto_id": produto.get("tiny_produto_id"),
                "sku": produto.get("sku"),
                "codigo": produto.get("codigo"),
                "descricao": produto.get("descricao"),
                "ncm": ncm_produto or None,
                "status": status,
                "percentual_confianca": confianca,
                "justificativa": justificativa,
                "revisao_manual": bool(regra.get("revisao_manual")) if regra else False,
                "data_auditoria": agora,
                "updated_at": agora,
            }
        )

        produto_atualizado = dict(produto)
        produto_atualizado.pop("status_tributario", None)
        produto_atualizado["status_ia"] = {
            "ok": "CONCLUIDO",
            "revisar": "REVISAO_MANUAL",
            "sem_regra": "PENDENTE",
            "erro": "ERRO",
        }[status]
        produto_atualizado["percentual_confianca"] = confianca
        produto_atualizado["justificativa_ia"] = justificativa
        produto_atualizado["data_analise"] = agora
        produto_atualizado["updated_at"] = agora
        produtos_atualizados.append(produto_atualizado)

    auditorias_salvas = _salvar_auditorias(auditorias)

    produtos_salvos = 0
    for inicio in range(0, len(produtos_atualizados), 100):
        bloco = produtos_atualizados[inicio: inicio + 100]
        salvos = supabase_upsert(bloco)
        produtos_salvos += len(salvos) if salvos else len(bloco)

    confianca_media = round(sum(confiancas) / len(confiancas), 2) if confiancas else 0.0

    return {
        "status": "ok",
        "filial": filial_normalizada,
        "contexto": {
            "regime_tributario": safe_str(regime_tributario).lower(),
            "uf_origem": origem,
            "uf_destino": destino,
            "tipo_operacao": safe_str(tipo_operacao).lower(),
            "finalidade": safe_str(finalidade).lower(),
            "consumidor_final": consumidor_final,
            "contribuinte_icms": contribuinte_icms,
            "data_referencia": referencia.isoformat(),
        },
        "total_produtos": len(produtos),
        "produtos_corretos": contadores["ok"],
        "produtos_revisao": contadores["revisar"],
        "produtos_sem_regra": contadores["sem_regra"],
        "produtos_com_erro": contadores["erro"],
        "confiabilidade_media": confianca_media,
        "regras_carregadas": len(regras),
        "auditorias_salvas": auditorias_salvas,
        "produtos_atualizados": produtos_salvos,
    }


def listar_produtos_auditoria(
    filial: str,
    status: Optional[str] = None,
    busca: Optional[str] = None,
    limite: int = 100,
    offset: int = 0,
    apenas_ultima: bool = True,
) -> List[Dict[str, Any]]:
    if limite < 1 or limite > 10000:
        raise HTTPException(status_code=400, detail="limite deve estar entre 1 e 10000.")
    if offset < 0:
        raise HTTPException(status_code=400, detail="offset não pode ser negativo.")

    filial_normalizada = normalizar_filial(filial)
    params: List[Tuple[str, str]] = [
        ("select", "*"),
        ("order", "data_auditoria.desc"),
        ("limit", "10000" if apenas_ultima else str(limite)),
        ("offset", "0" if apenas_ultima else str(offset)),
    ]
    if filial_normalizada != "all":
        params.append(("filial", f"eq.{filial_normalizada}"))
    if status:
        status_normalizado = safe_str(status).lower()
        if status_normalizado not in {"ok", "revisar", "sem_regra", "erro"}:
            raise HTTPException(status_code=400, detail="status inválido.")
        params.append(("status", f"eq.{status_normalizado}"))
    if busca:
        termo = montar_filtro_busca(busca)
        if termo:
            params.append(
                (
                    "or",
                    f"(descricao.ilike.*{termo}*,sku.ilike.*{termo}*,codigo.ilike.*{termo}*,ncm.ilike.*{termo}*)",
                )
            )

    dados = _supabase_request_tabela("GET", TABELA_AUDITORIA, params=params)
    registros = dados if isinstance(dados, list) else []

    if apenas_ultima:
        vistos = set()
        ultimos: List[Dict[str, Any]] = []
        for registro in registros:
            chave = registro.get("produto_id") or (
                registro.get("filial"), registro.get("tiny_produto_id")
            )
            if chave in vistos:
                continue
            vistos.add(chave)
            ultimos.append(registro)
        return ultimos[offset: offset + limite]

    return registros


def obter_resumo_auditoria(filial: str) -> Dict[str, Any]:
    registros = listar_produtos_auditoria(
        filial=filial,
        status=None,
        busca=None,
        limite=10000,
        offset=0,
        apenas_ultima=True,
    )

    contadores = {"ok": 0, "revisar": 0, "sem_regra": 0, "erro": 0}
    confiancas: List[float] = []
    ultima_auditoria: Optional[str] = None

    for registro in registros:
        status = safe_str(registro.get("status")).lower()
        if status in contadores:
            contadores[status] += 1
        confianca = _valor_float_seguro(registro.get("percentual_confianca"))
        if confianca > 0:
            confiancas.append(confianca)
        data_auditoria = safe_str(registro.get("data_auditoria"))
        if data_auditoria and (ultima_auditoria is None or data_auditoria > ultima_auditoria):
            ultima_auditoria = data_auditoria

    return {
        "status": "ok",
        "filial": normalizar_filial(filial),
        "total_produtos_auditados": len(registros),
        "produtos_corretos": contadores["ok"],
        "produtos_revisao": contadores["revisar"],
        "produtos_sem_regra": contadores["sem_regra"],
        "produtos_com_erro": contadores["erro"],
        "confiabilidade_media": (
            round(sum(confiancas) / len(confiancas), 2) if confiancas else 0.0
        ),
        "ultima_auditoria": ultima_auditoria,
    }


# ============================================================
# NCMs UTILIZADOS NOS PRODUTOS
# ============================================================

def listar_ncms_utilizados(
    filial: str,
    incluir_sem_ncm: bool = False,
    limite: int = 1000,
) -> Dict[str, Any]:
    """
    Lê todos os produtos da filial, agrupa por NCM e informa se existe
    regra tributária ativa cadastrada para cada código.

    O parâmetro limite restringe a quantidade de NCMs devolvidos, e não
    a quantidade de produtos lidos.
    """
    if limite < 1 or limite > 10000:
        raise HTTPException(
            status_code=400,
            detail="limite deve estar entre 1 e 10000.",
        )

    filial_normalizada = normalizar_filial(filial)

    params_produtos: List[Tuple[str, str]] = [
        ("select", "id,filial,ncm,descricao,sku,codigo,ativo,preco,custo,fornecedor"),
        ("order", "id.asc"),
    ]
    if filial_normalizada != "all":
        params_produtos.append(("filial", f"eq.{filial_normalizada}"))

    produtos = supabase_get_todos(
        params_produtos,
        tamanho_pagina=1000,
    )

    regras_ativas = _supabase_request_tabela(
        "GET",
        TABELA_REGRAS,
        params=[
            ("select", "ncm,id,status,percentual_confianca"),
            ("ativo", "eq.true"),
            ("limit", "10000"),
        ],
    )

    regras_por_ncm: Dict[str, List[Dict[str, Any]]] = {}
    for regra in regras_ativas or []:
        ncm_regra = ncm_limpo(regra.get("ncm"))
        if len(ncm_regra) == 8:
            regras_por_ncm.setdefault(ncm_regra, []).append(regra)

    agrupados: Dict[str, Dict[str, Any]] = {}
    produtos_sem_ncm = 0
    produtos_ncm_invalido = 0

    for produto in produtos:
        ncm = ncm_limpo(produto.get("ncm"))

        if not ncm:
            produtos_sem_ncm += 1
            if not incluir_sem_ncm:
                continue
            chave = "SEM_NCM"
            status = "sem_ncm"
        elif len(ncm) != 8:
            produtos_ncm_invalido += 1
            if not incluir_sem_ncm:
                continue
            chave = ncm
            status = "ncm_invalido"
        else:
            chave = ncm
            status = "valido"

        item = agrupados.setdefault(
            chave,
            {
                "ncm": None if chave == "SEM_NCM" else chave,
                "status_ncm": status,
                "quantidade_produtos": 0,
                "produtos_ativos": 0,
                "regra_cadastrada": False,
                "quantidade_regras_ativas": 0,
                "fornecedores_distintos": set(),
                "exemplos_produtos": [],
            },
        )

        item["quantidade_produtos"] += 1
        if produto.get("ativo") is True:
            item["produtos_ativos"] += 1

        fornecedor = safe_str(produto.get("fornecedor"))
        if fornecedor:
            item["fornecedores_distintos"].add(fornecedor)

        if len(item["exemplos_produtos"]) < 3:
            item["exemplos_produtos"].append(
                {
                    "descricao": produto.get("descricao"),
                    "sku": produto.get("sku"),
                    "codigo": produto.get("codigo"),
                }
            )

    for chave, item in agrupados.items():
        if chave != "SEM_NCM" and len(chave) == 8:
            regras = regras_por_ncm.get(chave, [])
            item["regra_cadastrada"] = bool(regras)
            item["quantidade_regras_ativas"] = len(regras)

        item["fornecedores_distintos"] = len(item["fornecedores_distintos"])

    lista_completa = sorted(
        agrupados.values(),
        key=lambda item: (-item["quantidade_produtos"], item.get("ncm") or ""),
    )
    lista = lista_completa[:limite]

    total_com_regra = sum(
        item["quantidade_produtos"]
        for item in lista_completa
        if item["regra_cadastrada"]
    )
    total_sem_regra = sum(
        item["quantidade_produtos"]
        for item in lista_completa
        if item["status_ncm"] == "valido" and not item["regra_cadastrada"]
    )

    return {
        "status": "ok",
        "filial": filial_normalizada,
        "total_produtos_lidos": len(produtos),
        "total_ncms_distintos": sum(
            1 for item in lista_completa if item["status_ncm"] == "valido"
        ),
        "total_ncms_retornados": len(lista),
        "produtos_com_regra_cadastrada": total_com_regra,
        "produtos_sem_regra_cadastrada": total_sem_regra,
        "produtos_sem_ncm": produtos_sem_ncm,
        "produtos_com_ncm_invalido": produtos_ncm_invalido,
        "ncms": lista,
    }
